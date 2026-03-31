import pandas as pd
from openpyxl import load_workbook
import json

file_path = r"D:\vk_work_dir\.vibe-kanban-workspaces\a977-brainstorming-xl\customer_platform_vk\.vibe-attachments\9f86035e-3114-4ef4-9436-75f56cdab6df_data.xlsx"

# 使用 openpyxl 直接读取
wb = load_workbook(file_path, data_only=True)


def decode_chinese(val):
    """尝试解码中文字符"""
    if val is None:
        return None
    if isinstance(val, str):
        # 尝试使用 gbk 解码
        try:
            # 先编码为 latin-1 再解码为 gbk
            return val.encode("latin-1").decode("gbk")
        except:
            pass
        try:
            # 如果已经是 unicode 但显示异常，直接返回
            return val
        except:
            pass
    return val


def decode_row(row):
    """解码一行数据"""
    return tuple(decode_chinese(v) for v in row)


# 分析所有 sheet
all_sheets_info = {}

for ws in wb.worksheets:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) > 0 and any(rows[0]):
        sheet_name = decode_chinese(ws.title) or ws.title
        decoded_cols = [decode_chinese(col) for col in rows[0]]

        all_sheets_info[sheet_name] = {
            "columns": decoded_cols,
            "row_count": len(rows) - 1,
            "sample_data": [decode_row(row) for row in rows[1:4]],
        }

        print(f"=== Sheet: {sheet_name} ===")
        print(f"行数：{len(rows) - 1}")
        print("列名:")
        for i, col in enumerate(decoded_cols):
            print(f"  {i + 1}. {col}")
        print()
        if len(rows) > 1:
            print("数据示例 (前 2 行):")
            for row in rows[1:3]:
                decoded = decode_row(row)
                print(f"  {decoded}")
        print()

# 保存信息到 JSON 文件
with open("excel_columns_info.json", "w", encoding="utf-8") as f:
    # 自定义 JSON 编码器处理 None 和特殊值
    def clean_for_json(obj):
        if obj is None:
            return None
        if isinstance(obj, dict):
            return {k: clean_for_json(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [clean_for_json(v) for v in obj]
        if isinstance(obj, tuple):
            return [clean_for_json(v) for v in obj]
        return obj

    json.dump(clean_for_json(all_sheets_info), f, ensure_ascii=False, indent=2)

print("列信息已保存到 excel_columns_info.json")
