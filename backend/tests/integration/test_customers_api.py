"""
Customers API 集成测试

测试覆盖：
1. GET /api/v1/customers - 客户列表（带筛选）
2. GET /api/v1/customers/:id - 客户详情
3. POST /api/v1/customers - 创建客户
4. PUT /api/v1/customers/:id - 更新客户
5. DELETE /api/v1/customers/:id - 删除客户
6. POST /api/v1/customers/import - Excel 导入
7. GET /api/v1/customers/export - Excel 导出
8. GET /api/v1/customers/:id/profile - 获取客户画像
9. PUT /api/v1/customers/:id/profile - 更新客户画像
10. GET /api/v1/customers/import-template - 下载导入模板
"""

import pytest
from sqlalchemy import text
import bcrypt
import io


@pytest.fixture
async def auth_token(test_client, test_user):
    """获取认证 Token"""
    login_request, login_response = await test_client.post(
        "/api/v1/auth/login",
        json={"username": test_user["username"], "password": test_user["password"]},
    )
    assert login_response.status == 200
    return login_response.json["data"]["access_token"]


@pytest.fixture
async def auth_headers(auth_token):
    """获取认证请求头"""
    return {"Authorization": f"Bearer {auth_token}"}


@pytest.fixture
def customer_data(db_session):
    """创建测试客户数据"""
    db_session.execute(text("DELETE FROM customers WHERE company_id LIKE 'TEST_%'"))
    db_session.commit()

    customers = [
        {
            "company_id": 1001,
            "name": "测试公司 1",
            "account_type": "正式账号",
            "customer_level": "KA",
            "settlement_type": "prepaid",
            "is_key_customer": True,
            "email": "test1@example.com",
        },
        {
            "company_id": 1002,
            "name": "测试公司 2",
            "account_type": "试用账号",
            "customer_level": "SMB",
            "settlement_type": "postpaid",
            "is_key_customer": False,
            "email": "test2@example.com",
        },
        {
            "company_id": 1003,
            "name": "测试公司 3",
            "account_type": "正式账号",
            "customer_level": "KA",
            "settlement_type": "prepaid",
            "is_key_customer": True,
            "email": "test3@example.com",
        },
    ]

    for cust in customers:
        db_session.execute(
            text(
                """
            INSERT INTO customers (company_id, name, account_type,
                customer_level, settlement_type, is_key_customer, email, created_at)
            VALUES (:company_id, :name, :account_type, :customer_level,
                :settlement_type, :is_key_customer, :email, NOW())
            """
            ),
            cust,
        )

    db_session.commit()

    result = db_session.execute(text("SELECT id FROM customers WHERE company_id = 1001"))
    customer_id = result.scalar_one()

    yield {"customers": customers, "customer_id": customer_id}

    db_session.execute(text("DELETE FROM customers WHERE company_id LIKE 'TEST_%'"))
    db_session.commit()


@pytest.mark.asyncio
async def test_list_customers_success(test_client, auth_headers, customer_data):
    """测试获取客户列表 - 成功场景"""
    request, response = await test_client.get(
        "/api/v1/customers",
        headers=auth_headers,
    )

    assert response.status == 200
    data = response.json
    assert data["code"] == 0
    assert data["message"] == "success"
    assert "list" in data["data"]
    assert "total" in data["data"]
    assert len(data["data"]["list"]) > 0


@pytest.mark.asyncio
async def test_list_customers_with_filters(test_client, auth_headers, customer_data):
    """测试获取客户列表 - 带筛选条件"""
    request, response = await test_client.get(
        "/api/v1/customers?customer_level=KA&is_key_customer=true",
        headers=auth_headers,
    )

    assert response.status == 200
    data = response.json
    assert data["code"] == 0
    assert len(data["data"]["list"]) >= 2

    for item in data["data"]["list"]:
        assert item["customer_level"] == "KA"
        assert item["is_key_customer"] is True


@pytest.mark.asyncio
async def test_list_customers_pagination(test_client, auth_headers, customer_data):
    """测试获取客户列表 - 分页"""
    request, response = await test_client.get(
        "/api/v1/customers?page=1&page_size=2",
        headers=auth_headers,
    )

    assert response.status == 200
    data = response.json
    assert data["code"] == 0
    assert data["data"]["page"] == 1
    assert data["data"]["page_size"] == 2
    assert len(data["data"]["list"]) <= 2


@pytest.mark.asyncio
async def test_get_customer_success(test_client, auth_headers, customer_data):
    """测试获取客户详情 - 成功场景"""
    customer_id = customer_data["customer_id"]

    request, response = await test_client.get(
        f"/api/v1/customers/{customer_id}",
        headers=auth_headers,
    )

    assert response.status == 200
    data = response.json
    assert data["code"] == 0
    assert data["data"]["company_id"] == 1001
    assert data["data"]["name"] == "测试公司 1"


@pytest.mark.asyncio
async def test_get_customer_not_found(test_client, auth_headers):
    """测试获取客户详情 - 客户不存在"""
    request, response = await test_client.get(
        "/api/v1/customers/999999",
        headers=auth_headers,
    )

    assert response.status == 404
    data = response.json
    assert data["code"] == 40401
    assert data["message"] == "客户不存在"


@pytest.mark.asyncio
async def test_create_customer_success(test_client, auth_headers, db_session):
    """测试创建客户 - 成功场景"""
    new_customer = {
        "company_id": 1000100,
        "name": "新创建测试公司",
        "account_type": "正式账号",
        "customer_level": "KA",
        "settlement_type": "prepaid",
        "is_key_customer": True,
        "email": "create_test@example.com",
    }

    request, response = await test_client.post(
        "/api/v1/customers",
        headers=auth_headers,
        json=new_customer,
    )

    assert response.status == 201
    data = response.json
    assert data["code"] == 0
    assert data["message"] == "创建成功"
    assert data["data"]["company_id"] == 1000100
    assert data["data"]["name"] == "新创建测试公司"

    db_session.execute(text("DELETE FROM customers WHERE company_id = 1000100"))
    db_session.commit()


@pytest.mark.asyncio
async def test_create_customer_missing_required_fields(test_client, auth_headers):
    """测试创建客户 - 缺少必填字段"""
    new_customer = {
        "name": "缺少公司 ID",
    }

    request, response = await test_client.post(
        "/api/v1/customers",
        headers=auth_headers,
        json=new_customer,
    )

    assert response.status == 400
    data = response.json
    assert data["code"] == 40001
    assert "公司 ID 和客户名称不能为空" in data["message"]


@pytest.mark.asyncio
async def test_create_customer_invalid_email(test_client, auth_headers):
    """测试创建客户 - 邮箱格式不正确"""
    new_customer = {
        "company_id": 999999,
        "name": "无效邮箱测试",
        "email": "invalid-email",
    }

    request, response = await test_client.post(
        "/api/v1/customers",
        headers=auth_headers,
        json=new_customer,
    )

    assert response.status == 400
    data = response.json
    assert data["code"] == 40002
    assert "邮箱格式不正确" in data["message"]


@pytest.mark.asyncio
async def test_update_customer_success(test_client, auth_headers, customer_data, db_session):
    """测试更新客户 - 成功场景"""
    customer_id = customer_data["customer_id"]

    update_data = {
        "name": "更新后的公司名称",
        "customer_level": "SMB",
        "is_key_customer": False,
    }

    request, response = await test_client.put(
        f"/api/v1/customers/{customer_id}",
        headers=auth_headers,
        json=update_data,
    )

    assert response.status == 200
    data = response.json
    assert data["code"] == 0
    assert data["message"] == "更新成功"

    result = db_session.execute(
        text("SELECT name, customer_level, is_key_customer FROM customers WHERE id = :id"),
        {"id": customer_id},
    )
    updated = result.fetchone()
    assert updated[0] == "更新后的公司名称"
    assert updated[1] == "SMB"
    assert updated[2] is False


@pytest.mark.asyncio
async def test_update_customer_not_found(test_client, auth_headers):
    """测试更新客户 - 客户不存在"""
    update_data = {"name": "不存在的客户"}

    request, response = await test_client.put(
        "/api/v1/customers/999999",
        headers=auth_headers,
        json=update_data,
    )

    assert response.status == 404
    data = response.json
    assert data["code"] == 40401
    assert data["message"] == "客户不存在"


@pytest.mark.asyncio
async def test_delete_customer_success(test_client, auth_headers, customer_data, db_session):
    """测试删除客户 - 成功场景"""
    customer_id = customer_data["customer_id"]

    request, response = await test_client.delete(
        f"/api/v1/customers/{customer_id}",
        headers=auth_headers,
    )

    assert response.status == 200
    data = response.json
    assert data["code"] == 0
    assert data["message"] == "删除成功"

    result = db_session.execute(
        text("SELECT deleted_at FROM customers WHERE id = :id"),
        {"id": customer_id},
    )
    deleted_at = result.scalar_one()
    assert deleted_at is not None


@pytest.mark.asyncio
async def test_delete_customer_not_found(test_client, auth_headers):
    """测试删除客户 - 客户不存在"""
    request, response = await test_client.delete(
        "/api/v1/customers/999999",
        headers=auth_headers,
    )

    assert response.status == 404
    data = response.json
    assert data["code"] == 40401
    assert data["message"] == "客户不存在"


@pytest.mark.asyncio
async def test_get_customer_profile_success(test_client, auth_headers, customer_data):
    """测试获取客户画像 - 成功场景"""
    customer_id = customer_data["customer_id"]

    request, response = await test_client.get(
        f"/api/v1/customers/{customer_id}/profile",
        headers=auth_headers,
    )

    assert response.status == 200
    data = response.json
    assert data["code"] == 0


@pytest.mark.asyncio
async def test_get_customer_profile_not_found(test_client, auth_headers):
    """测试获取客户画像 - 客户不存在"""
    request, response = await test_client.get(
        "/api/v1/customers/999999/profile",
        headers=auth_headers,
    )

    assert response.status == 404
    data = response.json
    assert data["code"] == 40401
    assert data["message"] == "客户不存在"


@pytest.mark.asyncio
async def test_update_customer_profile_success(
    test_client, auth_headers, customer_data, db_session
):
    """测试更新客户画像 - 成功场景"""
    customer_id = customer_data["customer_id"]

    profile_data = {
        "scale_level": "large",
        "consume_level": "high",
        "industry": "互联网",
        "is_real_estate": False,
        "description": "测试描述",
    }

    request, response = await test_client.put(
        f"/api/v1/customers/{customer_id}/profile",
        headers=auth_headers,
        json=profile_data,
    )

    assert response.status == 200
    data = response.json
    assert data["code"] == 0
    assert data["message"] == "更新成功"
    assert data["data"]["scale_level"] == "large"
    assert data["data"]["industry"] == "互联网"


@pytest.mark.asyncio
async def test_update_customer_profile_not_found(test_client, auth_headers):
    """测试更新客户画像 - 客户不存在"""
    profile_data = {"scale_level": "large"}

    request, response = await test_client.put(
        "/api/v1/customers/999999/profile",
        headers=auth_headers,
        json=profile_data,
    )

    assert response.status == 404
    data = response.json
    assert data["code"] == 40401
    assert data["message"] == "客户不存在"


@pytest.mark.asyncio
async def test_download_import_template(test_client, auth_headers):
    """测试下载 Excel 导入模板"""
    request, response = await test_client.get(
        "/api/v1/customers/import-template",
        headers=auth_headers,
    )

    assert response.status == 200
    assert (
        response.headers.get("Content-Type")
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "客户导入模板.xlsx" in response.headers.get("Content-Disposition", "")


@pytest.mark.asyncio
async def test_download_import_template_field_structure(test_client, auth_headers):
    """测试下载导入模板 - 验证模板字段结构正确性"""
    from openpyxl import load_workbook

    request, response = await test_client.get(
        "/api/v1/customers/import-template",
        headers=auth_headers,
    )

    assert response.status == 200

    # 解析返回的 Excel 文件
    wb = load_workbook(io.BytesIO(response.body))
    ws = wb.active

    assert ws.title == "客户导入模板"

    # 验证第 1 行：表头字段
    headers = [cell.value for cell in ws[1]]
    expected_headers = [
        "company_id",
        "name",
        "account_type",
        "industry",
        "customer_level",
        "price_policy",
        "settlement_cycle",
        "settlement_type",
        "is_key_customer",
        "email",
    ]
    assert headers == expected_headers, f"表头不匹配: {headers}"

    # 验证第 2 行：中文说明
    notes = [cell.value for cell in ws[2]]
    assert notes[0] == "必填"  # company_id
    assert notes[1] == "必填"  # name
    assert notes[2] == "可选"  # account_type
    assert notes[5] == "可选：定价/阶梯/包年"  # price_policy
    assert notes[7] == "可选：prepaid/postpaid"  # settlement_type
    assert notes[8] == "可选：true/false"  # is_key_customer

    # 验证第 3 行：示例数据
    example = [cell.value for cell in ws[3]]
    assert example[0] == 1001  # company_id 示例
    assert "示例公司" in str(example[1])  # name 示例
    assert example[5] == "定价"  # price_policy 示例
    assert example[7] == "prepaid"  # settlement_type 示例
    assert example[8] == "false"  # is_key_customer 示例
    assert "@" in str(example[9])  # email 示例


@pytest.mark.asyncio
async def test_download_import_template_header_count(test_client, auth_headers):
    """测试下载导入模板 - 验证表头数量为 10 个字段"""
    from openpyxl import load_workbook

    request, response = await test_client.get(
        "/api/v1/customers/import-template",
        headers=auth_headers,
    )

    assert response.status == 200

    wb = load_workbook(io.BytesIO(response.body))
    ws = wb.active

    header_count = sum(1 for cell in ws[1] if cell.value is not None)
    assert header_count == 10, f"期望 10 个表头，实际 {header_count} 个"


@pytest.mark.asyncio
async def test_download_import_template_unauthorized(test_client):
    """测试下载导入模板 - 未认证访问应返回 401"""
    request, response = await test_client.get(
        "/api/v1/customers/import-template",
    )

    assert response.status == 401


@pytest.mark.asyncio
async def test_download_import_template_no_import_permission(test_client, db_session):
    """测试下载导入模板 - 无 import 权限的用户仍可下载（权限已放开）"""

    import time

    unique_suffix = int(time.time())
    username = f"template_only_user_{unique_suffix}"
    role_name = f"view_only_{unique_suffix}"
    password = "test123456"

    # 清理旧数据
    db_session.execute(
        text("DELETE FROM users WHERE username = :username"),
        {"username": username},
    )
    db_session.commit()

    password_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    db_session.execute(
        text(
            """
        INSERT INTO users (username, password_hash, email, is_active, created_at)
        VALUES (:username, :password_hash, :email, :is_active, NOW())
        """
        ),
        {
            "username": username,
            "password_hash": password_hash,
            "email": "template_only@example.com",
            "is_active": True,
        },
    )

    # 创建只有 customers:view 权限的角色
    db_session.execute(text("DELETE FROM roles WHERE name = :role_name"), {"role_name": role_name})
    db_session.execute(
        text(
            """
        INSERT INTO roles (name, description, created_at)
        VALUES (:name, :description, NOW())
        """
        ),
        {"name": role_name, "description": "仅查看角色"},
    )
    result = db_session.execute(
        text("SELECT id FROM roles WHERE name = :role_name"), {"role_name": role_name}
    )
    role_id = result.scalar_one()

    # 仅赋予 customers:view 权限（不赋予 customers:import）
    # 确保权限存在（test_user fixture 未使用时权限表可能为空）
    result = db_session.execute(text("SELECT id FROM permissions WHERE code = 'customers:view'"))
    perm_row = result.fetchone()
    if perm_row is None:
        db_session.execute(
            text(
                """
            INSERT INTO permissions (code, name, description, module, created_at)
            VALUES ('customers:view', '查看客户', '查看客户列表和详情', 'customers', NOW())
            """
            )
        )
        db_session.commit()
        result = db_session.execute(
            text("SELECT id FROM permissions WHERE code = 'customers:view'")
        )
    perm_id = result.scalar_one()

    db_session.execute(
        text("DELETE FROM role_permissions WHERE role_id = :role_id"),
        {"role_id": role_id},
    )
    db_session.execute(
        text(
            """
        INSERT INTO role_permissions (role_id, permission_id)
        VALUES (:role_id, :permission_id)
        """
        ),
        {"role_id": role_id, "permission_id": perm_id},
    )

    # 关联用户到角色
    result = db_session.execute(
        text("SELECT id FROM users WHERE username = :username"),
        {"username": username},
    )
    user_id = result.scalar_one()

    db_session.execute(
        text("DELETE FROM user_roles WHERE user_id = :user_id"),
        {"user_id": user_id},
    )
    db_session.execute(
        text(
            """
        INSERT INTO user_roles (user_id, role_id)
        VALUES (:user_id, :role_id)
        """
        ),
        {"user_id": user_id, "role_id": role_id},
    )
    db_session.commit()

    try:
        # 登录
        login_request, login_response = await test_client.post(
            "/api/v1/auth/login",
            json={"username": username, "password": password},
        )
        assert login_response.status == 200
        token = login_response.json["data"]["access_token"]

        # 下载模板 - 应该成功（200），因为不再需要 customers:import 权限
        request, response = await test_client.get(
            "/api/v1/customers/import-template",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status == 200
        assert (
            response.headers.get("Content-Type")
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    finally:
        db_session.execute(
            text(
                "DELETE FROM user_roles WHERE user_id = (SELECT id FROM users WHERE username = :username)"
            ),
            {"username": username},
        )
        db_session.execute(
            text(
                "DELETE FROM role_permissions WHERE role_id = (SELECT id FROM roles WHERE name = :role_name)"
            ),
            {"role_name": role_name},
        )
        db_session.execute(
            text("DELETE FROM users WHERE username = :username"),
            {"username": username},
        )
        db_session.execute(
            text("DELETE FROM roles WHERE name = :role_name"),
            {"role_name": role_name},
        )
        db_session.commit()


@pytest.mark.asyncio
async def test_import_customers_success(test_client, auth_headers, db_session):
    """测试 Excel 导入客户 - 成功场景"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "company_id",
            "name",
            "account_type",
            "industry",
            "customer_level",
            "settlement_type",
            "is_key_customer",
            "email",
        ]
    )
    ws.append(
        [
            1000001,
            "导入测试公司 1",
            "正式账号",
            "互联网",
            "KA",
            "prepaid",
            "false",
            "import1@example.com",
        ]
    )
    ws.append(
        [
            1000002,
            "导入测试公司 2",
            "试用账号",
            "房地产",
            "SMB",
            "postpaid",
            "true",
            "import2@example.com",
        ]
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    files = {
        "file": (
            "test_import.xlsx",
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }

    request, response = await test_client.post(
        "/api/v1/customers/import",
        headers=auth_headers,
        files=files,
    )

    assert response.status == 200
    data = response.json
    assert data["code"] == 0
    assert data["message"] == "导入完成"
    assert data["data"]["success_count"] == 2

    db_session.execute(
        text("DELETE FROM customers WHERE company_id >= 1000000 AND company_id < 1000010")
    )
    db_session.commit()


@pytest.mark.asyncio
async def test_import_customers_missing_file(test_client, auth_headers):
    """测试 Excel 导入客户 - 缺少文件"""
    request, response = await test_client.post(
        "/api/v1/customers/import",
        headers=auth_headers,
    )

    assert response.status == 400
    data = response.json
    assert data["code"] == 40001
    assert data["message"] == "请上传 Excel 文件"


@pytest.mark.asyncio
async def test_import_customers_wrong_format(test_client, auth_headers):
    """测试 Excel 导入客户 - 文件格式错误"""
    files = {"file": ("test.txt", b"not an excel file", "text/plain")}

    request, response = await test_client.post(
        "/api/v1/customers/import",
        headers=auth_headers,
        files=files,
    )

    assert response.status == 400
    data = response.json
    assert data["code"] == 40002
    assert data["message"] == "请上传 .xlsx 格式的文件"


@pytest.mark.asyncio
async def test_import_customers_missing_columns(test_client, auth_headers):
    """测试 Excel 导入客户 - 缺少必填列"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["company_id"])
    ws.append(["MISSING_NAME"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    files = {
        "file": (
            "test_missing.xlsx",
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }

    request, response = await test_client.post(
        "/api/v1/customers/import",
        headers=auth_headers,
        files=files,
    )

    assert response.status == 400
    data = response.json
    assert data["code"] == 40003
    assert "缺少必填列" in data["message"]


@pytest.mark.asyncio
async def test_import_customers_with_template_notes_row(test_client, auth_headers, db_session):
    """测试导入带中文说明行的模板文件（智能跳过逻辑）"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    # 第 1 行：列名
    ws.append(["company_id", "name", "account_type"])
    # 第 2 行：中文说明（模板特征）
    ws.append(["必填", "必填", "可选"])
    # 第 3 行：实际数据
    ws.append([1000010, "模板测试公司", "正式账号"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    files = {
        "file": (
            "test_template.xlsx",
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }

    request, response = await test_client.post(
        "/api/v1/customers/import",
        headers=auth_headers,
        files=files,
    )

    assert response.status == 200
    data = response.json
    assert data["code"] == 0
    assert data["data"]["success_count"] == 1

    # 清理测试数据
    db_session.execute(
        text("DELETE FROM customers WHERE company_id >= 1000010 AND company_id < 1000020")
    )
    db_session.commit()


@pytest.mark.asyncio
async def test_import_customers_invalid_email_should_fail(test_client, auth_headers, db_session):
    """测试导入 - 邮箱格式错误应返回错误信息"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["company_id", "name", "email"])
    ws.append([1000020, "邮箱错误测试", "not-an-email"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    files = {
        "file": (
            "test_bad_email.xlsx",
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }

    request, response = await test_client.post(
        "/api/v1/customers/import",
        headers=auth_headers,
        files=files,
    )

    assert response.status == 200
    data = response.json
    assert data["data"]["error_count"] >= 1

    db_session.execute(
        text("DELETE FROM customers WHERE company_id >= 1000020 AND company_id < 1000030")
    )
    db_session.commit()


@pytest.mark.asyncio
async def test_import_customers_empty_required_fields(test_client, auth_headers, db_session):
    """测试导入 - 必填字段为空应返回错误"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["company_id", "name"])
    # Use None (NaN) values - pandas will include these rows
    ws.append([None, "Company with empty company_id"])
    ws.append(["EMPTY_NAME", None])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    files = {
        "file": (
            "test_empty_required.xlsx",
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }

    request, response = await test_client.post(
        "/api/v1/customers/import",
        headers=auth_headers,
        files=files,
    )

    assert response.status == 200
    data = response.json
    assert data["code"] == 0
    # Empty company_id and name should be recorded as errors
    assert data["data"]["error_count"] >= 2
    assert data["data"]["success_count"] == 0


@pytest.mark.asyncio
async def test_import_customers_duplicate_company_id(test_client, auth_headers, db_session):
    """测试导入 - 重复 company_id 应部分成功或报错"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["company_id", "name"])
    ws.append([1000030, "公司 A"])
    ws.append([1000030, "公司 B"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    files = {
        "file": (
            "test_duplicate.xlsx",
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }

    request, response = await test_client.post(
        "/api/v1/customers/import",
        headers=auth_headers,
        files=files,
    )

    assert response.status == 200
    data = response.json
    assert data["code"] == 0
    assert data["data"]["success_count"] >= 1 or data["data"]["error_count"] >= 1

    db_session.execute(
        text("DELETE FROM customers WHERE company_id >= 1000030 AND company_id < 1000040")
    )
    db_session.commit()


@pytest.mark.asyncio
async def test_import_customers_invalid_price_policy(test_client, auth_headers):
    """测试导入 - 非法 price_policy 值应被正确处理"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["company_id", "name", "price_policy"])
    ws.append(["INVALID_POLICY", "非法策略测试", "非法值"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    files = {
        "file": (
            "test_invalid_policy.xlsx",
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }

    request, response = await test_client.post(
        "/api/v1/customers/import",
        headers=auth_headers,
        files=files,
    )

    assert response.status == 200
    data = response.json
    assert data["code"] == 0
    assert data["data"]["error_count"] >= 1


@pytest.mark.asyncio
async def test_import_customers_without_notes_row(test_client, auth_headers, db_session):
    """测试导入普通用户文件（无说明行，不应跳过）"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["company_id", "name"])
    ws.append([1000040, "普通公司 A"])
    ws.append([1000041, "普通公司 B"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    files = {
        "file": (
            "test_normal.xlsx",
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }

    request, response = await test_client.post(
        "/api/v1/customers/import",
        headers=auth_headers,
        files=files,
    )

    assert response.status == 200
    data = response.json
    assert data["code"] == 0
    assert data["data"]["success_count"] == 2

    # 清理测试数据
    db_session.execute(
        text("DELETE FROM customers WHERE company_id >= 1000040 AND company_id < 1000050")
    )
    db_session.commit()


@pytest.mark.asyncio
async def test_export_customers_success(test_client, auth_headers, customer_data):
    """测试 Excel 导出客户 - 成功场景"""
    request, response = await test_client.get(
        "/api/v1/customers/export",
        headers=auth_headers,
    )

    assert response.status == 200
    assert (
        response.headers.get("Content-Type")
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "customers_" in response.headers.get("Content-Disposition", "")
    assert ".xlsx" in response.headers.get("Content-Disposition", "")


@pytest.mark.asyncio
async def test_export_customers_with_filters(test_client, auth_headers, customer_data):
    """测试 Excel 导出客户 - 带筛选条件"""
    request, response = await test_client.get(
        "/api/v1/customers/export?customer_level=KA",
        headers=auth_headers,
    )

    assert response.status == 200
    assert (
        response.headers.get("Content-Type")
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@pytest.mark.asyncio
async def test_export_customers_contains_test_data(test_client, auth_headers, customer_data):
    """测试导出 - 验证导出文件包含测试数据"""
    from openpyxl import load_workbook

    request, response = await test_client.get(
        "/api/v1/customers/export",
        headers=auth_headers,
    )

    assert response.status == 200

    wb = load_workbook(io.BytesIO(response.body))
    ws = wb.active

    row_count = ws.max_row
    assert row_count >= 2, f"期望至少 2 行，实际 {row_count} 行"

    found_test_data = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and "TEST" in str(row[0]):
            found_test_data = True
            break
    assert found_test_data, "导出文件未找到测试数据"


@pytest.mark.asyncio
async def test_export_customers_field_consistency(test_client, auth_headers, customer_data):
    """测试导出 - 验证导出文件字段与导入模板字段一致性"""
    from openpyxl import load_workbook

    request, response = await test_client.get(
        "/api/v1/customers/export",
        headers=auth_headers,
    )

    assert response.status == 200

    wb = load_workbook(io.BytesIO(response.body))
    ws = wb.active

    export_headers = [cell.value for cell in ws[1] if cell.value is not None]

    template_headers = [
        "company_id",
        "name",
        "account_type",
        "industry",
        "customer_level",
        "price_policy",
        "settlement_cycle",
        "settlement_type",
        "is_key_customer",
        "email",
    ]

    for header in template_headers:
        assert header in export_headers, f"导出文件缺少字段: {header}"

    # Verify company_id values are integers
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            assert isinstance(row[0], (int, float)), f"company_id 应为整数，实际为 {type(row[0])}"


@pytest.mark.asyncio
async def test_customers_unauthorized(test_client):
    """测试未认证访问"""
    request, response = await test_client.get("/api/v1/customers")

    assert response.status in [401, 403]


@pytest.mark.asyncio
async def test_customers_missing_permission(test_client, db_session):
    """测试缺少权限访问"""
    from unittest.mock import AsyncMock

    username = "no_perm_user"
    password = "test123456"
    import bcrypt

    db_session.execute(
        text("DELETE FROM users WHERE username = :username"),
        {"username": username},
    )
    password_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    db_session.execute(
        text(
            """
        INSERT INTO users (username, password_hash, email, is_active, created_at)
        VALUES (:username, :password_hash, :email, :is_active, NOW())
        """
        ),
        {
            "username": username,
            "password_hash": password_hash,
            "email": "noperm@example.com",
            "is_active": True,
        },
    )
    db_session.commit()

    try:
        login_request, login_response = await test_client.post(
            "/api/v1/auth/login",
            json={"username": username, "password": password},
        )
        assert login_response.status == 200
        token = login_response.json["data"]["access_token"]

        # Patch auth.py 中的权限缓存引用（装饰器持有的是这个模块的引用）
        from app.middleware import auth as auth_module

        original_get = auth_module.permission_cache.get_permissions
        auth_module.permission_cache.get_permissions = AsyncMock(return_value=set())

        try:
            request, response = await test_client.get(
                "/api/v1/customers",
                headers={"Authorization": f"Bearer {token}"},
            )

            assert response.status == 403
        finally:
            auth_module.permission_cache.get_permissions = original_get
    finally:
        db_session.execute(
            text("DELETE FROM users WHERE username = :username"),
            {"username": username},
        )
        db_session.commit()
