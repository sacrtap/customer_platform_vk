"""
数据清洗脚本：将 docs/data.xlsx 转换为符合导入模板格式的 clean 数据
输出: docs/data_clean.xlsx

字段映射规则：
- 公司id → company_id
- 公司名称 → name
- 账号类型 → account_type（正式→正式账号，客户测试账号→客户测试账号，众趣内部→内部账号）
- 行业类型 → industry（"无"→空）
- 结算方式 → price_policy（定价结算→定价，包年套餐→包年，易遨结算→定价）
- 结算方式 → settlement_cycle（定价结算→月结，包年套餐→年结）
- 结算方式 → settlement_type（统一 prepaid）
- 客户等级 → is_key_customer（S/A→true，其他→false）
- 所属ERP → erp_system
- 首次回款时间 → first_payment_date
- 接入时间 → onboarding_date（文字描述追加到 notes）
- 合作状态 → cooperation_status（正常使用→active，近一年未使用→noused，商务阶段→noused，待确认→noused）
- 是否结算 → is_settlement_enabled
- 是否停用 → is_disabled
- 备注 → notes
- 客户消费等级 → consume_level（C2→A, C3→B, C4→C, C5→D, C6→E）
- 月均拍摄量 → monthly_avg_shots
- 月均拍摄量（测算） → monthly_avg_shots_estimated
- 预估年消费 → estimated_annual_spend
- 25年实际消费 → actual_annual_spend_2025
"""
import openpyxl
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).parent.parent.parent
INPUT_FILE = PROJECT_ROOT / "docs" / "data.xlsx"
OUTPUT_FILE = PROJECT_ROOT / "docs" / "data_clean.xlsx"

# 输出字段列表
OUTPUT_FIELDS = [
    "company_id",
    "name",
    "account_type",
    "industry",
    "price_policy",
    "settlement_type",
    "settlement_cycle",
    "is_key_customer",
    "email",
    "erp_system",
    "first_payment_date",
    "onboarding_date",
    "cooperation_status",
    "is_settlement_enabled",
    "is_disabled",
    "notes",
    "scale_level",
    "consume_level",
    "monthly_avg_shots",
    "monthly_avg_shots_estimated",
    "estimated_annual_spend",
    "actual_annual_spend_2025",
]

# 账号类型映射
ACCOUNT_TYPE_MAP = {
    "正式": "正式账号",
    "客户测试账号": "客户测试账号",
    "众趣内部": "内部账号",
}

# 计费模式映射
PRICE_POLICY_MAP = {
    "定价结算": "定价",
    "包年套餐": "包年",
    "包年限量套餐": "包年",
    "易遨结算": "定价",
}

# 结算周期映射
SETTLEMENT_CYCLE_MAP = {
    "定价结算": "monthly",
    "包年套餐": "yearly",
    "包年限量套餐": "yearly",
    "易遨结算": "monthly",
}

# 合作状态映射
COOPERATION_STATUS_MAP = {
    "正常使用": "active",
    "近一年未使用": "noused",
    "商务阶段": "noused",
    "待确认": "noused",
    "终止": "terminated",
    "暂停": "suspended",
}

# 消费等级映射
CONSUME_LEVEL_MAP = {
    "C2": "A",
    "C3": "B",
    "C4": "C",
    "C5": "D",
    "C6": "E",
}


def clean_value(val, field_name):
    """清洗单个值"""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val in ("#N/A", "#VALUE!", "None", "", "nan"):
            return None
    
    # 账号类型
    if field_name == "account_type":
        return ACCOUNT_TYPE_MAP.get(str(val), str(val))
    
    # 行业类型："无"→空
    if field_name == "industry":
        if str(val) == "无":
            return None
        return str(val)
    
    # 计费模式
    if field_name == "price_policy":
        return PRICE_POLICY_MAP.get(str(val))
    
    # 结算周期
    if field_name == "settlement_cycle":
        return SETTLEMENT_CYCLE_MAP.get(str(val))
    
    # 结算方式
    if field_name == "settlement_type":
        return "prepaid"
    
    # 重点客户
    if field_name == "is_key_customer":
        return str(val) in ("S", "A")
    
    # 首次回款时间
    if field_name == "first_payment_date":
        if hasattr(val, "strftime"):
            return val.strftime("%Y-%m-%d")
        return None
    
    # 接入时间（文字描述，返回 None，由调用方处理追加到 notes）
    if field_name == "onboarding_date":
        if hasattr(val, "strftime"):
            return val.strftime("%Y-%m-%d")
        # 文字描述如 "25上半年" 不转为日期
        return None
    
    # 合作状态
    if field_name == "cooperation_status":
        return COOPERATION_STATUS_MAP.get(str(val))
    
    # 是否结算
    if field_name == "is_settlement_enabled":
        v = str(val).lower()
        if v in ("是", "true", "1", "yes"):
            return "true"
        if v in ("否", "false", "0", "no"):
            return "false"
        return None
    
    # 是否停用
    if field_name == "is_disabled":
        v = str(val).lower()
        if v in ("是", "true", "1", "yes"):
            return "true"
        if v in ("否", "false", "0", "no"):
            return "false"
        return None
    
    # 消费等级
    if field_name == "consume_level":
        v = str(val).strip()
        if v in ("0", "0.0"):
            return None
        return CONSUME_LEVEL_MAP.get(v)
    
    # 月均拍摄量（整数，0→空）
    if field_name in ("monthly_avg_shots", "monthly_avg_shots_estimated"):
        if val == 0 or val == "0":
            return None
        try:
            return int(float(str(val)))
        except (ValueError, TypeError):
            return None
    
    # 金额字段（0/#VALUE!→空）
    if field_name in ("estimated_annual_spend", "actual_annual_spend_2025"):
        if val == 0 or val == "0" or str(val) in ("#VALUE!", "#N/A"):
            return None
        try:
            return float(str(val))
        except (ValueError, TypeError):
            return None
    
    return val


def main():
    wb_in = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws_in = wb_in.active

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "客户导入模板"

    # 写入表头
    ws_out.append(OUTPUT_FIELDS)

    # 添加说明行
    notes_row = [
        "必填：唯一整数",
        "必填：客户名称",
        "可选：正式账号/客户测试账号/内部账号",
        "可选：行业类型",
        "可选：定价/阶梯/包年",
        "可选：prepaid=预付费, postpaid=后付费",
        "可选：日结/周结/月结/季结",
        "可选：true/false",
        "可选：邮箱地址",
        "可选：所属 ERP 系统",
        "可选：YYYY-MM-DD 格式",
        "可选：YYYY-MM-DD 格式",
        "可选：active=合作中, suspended=暂停, terminated=终止, noused=近一年未使用",
        "可选：true/false",
        "可选：true/false",
        "可选：备注信息",
        "可选：100/500/1000/2000/5000",
        "可选：S/A/B/C/D/E",
        "可选：整数",
        "可选：整数",
        "可选：金额",
        "可选：金额",
    ]
    ws_out.append(notes_row)

    # 读取输入数据
    input_headers = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    
    # 建立列索引
    col_indices = {}
    source_field_map = {
        "公司id": "company_id",
        "公司名称": "name",
        "账号类型": "account_type",
        "行业类型": "industry",
        "所属ERP": "erp_system",
        "首次回款时间": "first_payment_date",
        "接入时间": "onboarding_date_raw",
        "合作状态": "cooperation_status",
        "是否结算": "is_settlement_enabled",
        "是否停用": "is_disabled",
        "备注": "notes",
        "结算方式": "settlement_source",
        "客户消费等级": "consume_level",
        "月均拍摄量": "monthly_avg_shots",
        "月均拍摄量（测算）": "monthly_avg_shots_estimated",
        "预估年消费": "estimated_annual_spend",
        "25年实际消费": "actual_annual_spend_2025",
        "客户等级": "is_key_customer_source",
    }
    
    for cn_name, en_name in source_field_map.items():
        if cn_name in input_headers:
            col_indices[en_name] = input_headers.index(cn_name)

    cleaned_count = 0
    skipped = 0

    for row_idx, row in enumerate(ws_in.iter_rows(min_row=2, max_row=ws_in.max_row, values_only=True), 2):
        # 获取关键字段
        company_id = row[col_indices["company_id"]] if "company_id" in col_indices else None
        name = row[col_indices["name"]] if "name" in col_indices else None

        # 跳过无效行
        if company_id is None or (isinstance(company_id, float) and str(company_id) == "nan"):
            skipped += 1
            continue
        if name is None or (isinstance(name, float) and str(name) == "nan") or str(name).strip() == "":
            skipped += 1
            continue

        # 构建输出行
        output_row = []
        
        # 接入时间原文（用于追加到 notes）
        onboarding_raw = None
        if "onboarding_date_raw" in col_indices:
            idx = col_indices["onboarding_date_raw"]
            val = row[idx] if idx < len(row) else None
            if val and str(val).strip() not in ("", "#N/A", "None", "nan"):
                onboarding_raw = str(val).strip()
        
        # 备注原文
        notes_raw = None
        if "notes" in col_indices:
            idx = col_indices["notes"]
            val = row[idx] if idx < len(row) else None
            if val and str(val).strip() not in ("", "#N/A", "None", "nan"):
                notes_raw = str(val).strip()
        
        # 合并 notes
        final_notes = notes_raw
        if onboarding_raw:
            if final_notes:
                final_notes = f"{final_notes}；接入时间：{onboarding_raw}"
            else:
                final_notes = f"接入时间：{onboarding_raw}"

        for field in OUTPUT_FIELDS:
            if field == "notes":
                output_row.append(final_notes)
                continue
            if field == "onboarding_date":
                # 接入时间不填充（文字描述无法转为标准日期）
                output_row.append(None)
                continue
            if field == "scale_level":
                # 规模等级不映射
                output_row.append(None)
                continue
            if field == "email":
                # 源数据无邮箱
                output_row.append(None)
                continue
            
            # 获取源数据列
            source_col = None
            if field == "company_id":
                source_col = "company_id"
            elif field == "name":
                source_col = "name"
            elif field == "account_type":
                source_col = "account_type"
            elif field == "industry":
                source_col = "industry"
            elif field == "price_policy":
                source_col = "settlement_source"
            elif field == "settlement_cycle":
                source_col = "settlement_source"
            elif field == "settlement_type":
                source_col = "settlement_source"
            elif field == "is_key_customer":
                source_col = "is_key_customer_source"
            elif field == "erp_system":
                source_col = "erp_system"
            elif field == "first_payment_date":
                source_col = "first_payment_date"
            elif field == "cooperation_status":
                source_col = "cooperation_status"
            elif field == "is_settlement_enabled":
                source_col = "is_settlement_enabled"
            elif field == "is_disabled":
                source_col = "is_disabled"
            elif field == "consume_level":
                source_col = "consume_level"
            elif field == "monthly_avg_shots":
                source_col = "monthly_avg_shots"
            elif field == "monthly_avg_shots_estimated":
                source_col = "monthly_avg_shots_estimated"
            elif field == "estimated_annual_spend":
                source_col = "estimated_annual_spend"
            elif field == "actual_annual_spend_2025":
                source_col = "actual_annual_spend_2025"
            
            if source_col and source_col in col_indices:
                idx = col_indices[source_col]
                raw_val = row[idx] if idx < len(row) else None
                cleaned = clean_value(raw_val, field)
                output_row.append(cleaned)
            else:
                output_row.append(None)

        ws_out.append(output_row)
        cleaned_count += 1

    wb_out.save(OUTPUT_FILE)
    print(f"清洗完成: {cleaned_count} 行有效数据, {skipped} 行被跳过")
    print(f"输出文件: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
