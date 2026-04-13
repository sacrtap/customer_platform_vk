"""结算管理路由"""

from sanic import Blueprint
from sanic.response import json, file as response_file
from sanic.request import Request
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from ..services.billing import BalanceService, PricingService, InvoiceService
from ..middleware.auth import get_current_user, require_permission, auth_required
from ..cache.base import cache_service

billing_bp = Blueprint("billing", url_prefix="/api/v1/billing")


# ==================== 余额管理 ====================


@billing_bp.get("/balances")
@auth_required
@require_permission("billing:view")
async def get_balances(request: Request):
    """获取余额列表（支持服务端筛选和分页）"""
    db: AsyncSession = request.ctx.db_session

    from ..models.billing import CustomerBalance
    from sqlalchemy import select, func
    from sqlalchemy.orm import selectinload
    from ..models.customers import Customer

    # 分页参数
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 20))
    page_size = min(page_size, 100)

    # 筛选参数
    keyword = request.args.get("keyword")  # 客户名称模糊搜索
    customer_id = (
        int(request.args.get("customer_id"))
        if request.args.get("customer_id")
        else None
    )

    base_stmt = (
        select(CustomerBalance)
        .join(Customer, CustomerBalance.customer_id == Customer.id)
        .where(
            CustomerBalance.deleted_at.is_(None),
            Customer.deleted_at.is_(None),
        )
    )

    # 服务端过滤
    if customer_id:
        base_stmt = base_stmt.where(CustomerBalance.customer_id == customer_id)
    if keyword:
        base_stmt = base_stmt.where(Customer.name.ilike(f"%{keyword}%"))

    # 总数查询
    count_stmt = (
        select(func.count(CustomerBalance.id))
        .join(Customer, CustomerBalance.customer_id == Customer.id)
        .where(
            CustomerBalance.deleted_at.is_(None),
            Customer.deleted_at.is_(None),
        )
    )
    if customer_id:
        count_stmt = count_stmt.where(CustomerBalance.customer_id == customer_id)
    if keyword:
        count_stmt = count_stmt.where(Customer.name.ilike(f"%{keyword}%"))

    total = (await db.execute(count_stmt)).scalar()

    # 分页查询
    stmt = base_stmt.options(selectinload(CustomerBalance.customer))
    stmt = stmt.order_by(CustomerBalance.updated_at.desc())
    stmt = stmt.offset((page - 1) * page_size).limit(page_size)

    result = await db.execute(stmt)
    balances = result.scalars().all()

    return json(
        {
            "code": 0,
            "message": "success",
            "data": {
                "list": [
                    {
                        "id": b.id,
                        "customer_id": b.customer_id,
                        "customer_name": b.customer.name if b.customer else None,
                        "total_amount": float(b.total_amount) if b.total_amount else 0,
                        "real_amount": float(b.real_amount) if b.real_amount else 0,
                        "bonus_amount": float(b.bonus_amount) if b.bonus_amount else 0,
                        "used_total": float(b.used_total) if b.used_total else 0,
                        "used_real": float(b.used_real) if b.used_real else 0,
                        "used_bonus": float(b.used_bonus) if b.used_bonus else 0,
                    }
                    for b in balances
                ],
                "total": total,
                "page": page,
                "page_size": page_size,
            },
        }
    )


@billing_bp.get("/customers/<customer_id:int>/balance")
@auth_required
@require_permission("billing:view")
async def get_customer_balance(request: Request, customer_id: int):
    """获取客户余额"""
    db: AsyncSession = request.ctx.db_session
    balance_service = BalanceService(db)

    balance = await balance_service.get_balance_by_customer_id(customer_id)

    if not balance:
        return json(
            {
                "code": 0,
                "message": "success",
                "data": {
                    "total_amount": 0,
                    "real_amount": 0,
                    "bonus_amount": 0,
                    "used_total": 0,
                    "used_real": 0,
                    "used_bonus": 0,
                },
            }
        )

    return json(
        {
            "code": 0,
            "message": "success",
            "data": {
                "total_amount": float(balance.total_amount)
                if balance.total_amount
                else 0,
                "real_amount": float(balance.real_amount) if balance.real_amount else 0,
                "bonus_amount": float(balance.bonus_amount)
                if balance.bonus_amount
                else 0,
                "used_total": float(balance.used_total) if balance.used_total else 0,
                "used_real": float(balance.used_real) if balance.used_real else 0,
                "used_bonus": float(balance.used_bonus) if balance.used_bonus else 0,
            },
        }
    )


@billing_bp.post("/recharge")
@auth_required
@require_permission("billing:recharge")
async def recharge(request: Request):
    """
    客户充值

    Body:
    {
        "customer_id": 1,
        "real_amount": 10000.00,
        "bonus_amount": 2000.00,
        "payment_proof": "/uploads/proof.png",
        "remark": "Q1 季度充值"
    }
    """
    db: AsyncSession = request.ctx.db_session
    data = request.json
    user = get_current_user(request)

    customer_id = data.get("customer_id")
    real_amount = Decimal(str(data.get("real_amount", 0)))
    bonus_amount = Decimal(str(data.get("bonus_amount", 0)))

    if not customer_id or real_amount <= 0:
        return json(
            {"code": 40001, "message": "客户 ID 和实充金额不能为空"},
            status=400,
        )

    balance_service = BalanceService(db)

    record = await balance_service.recharge(
        customer_id=customer_id,
        real_amount=real_amount,
        bonus_amount=bonus_amount,
        operator_id=user["user_id"] if user else 1,
        payment_proof=data.get("payment_proof"),
        remark=data.get("remark"),
    )

    # 充值后清除相关缓存
    await cache_service.invalidate_analytics_cache("health")
    await cache_service.invalidate_analytics_cache("dashboard")
    await cache_service.invalidate_customer_cache(customer_id)

    return json(
        {
            "code": 0,
            "message": "充值成功",
            "data": {
                "id": record.id,
                "customer_id": record.customer_id,
                "real_amount": float(record.real_amount),
                "bonus_amount": float(record.bonus_amount),
                "total_amount": float(record.real_amount + record.bonus_amount),
            },
        },
        status=201,
    )


@billing_bp.get("/recharge-records")
@auth_required
@require_permission("billing:view")
async def get_recharge_records(request: Request):
    """获取充值记录列表"""
    db: AsyncSession = request.ctx.db_session
    balance_service = BalanceService(db)

    customer_id = (
        int(request.args.get("customer_id"))
        if request.args.get("customer_id")
        else None
    )
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 20))

    records, total = await balance_service.get_recharge_records(
        customer_id=customer_id,
        page=page,
        page_size=page_size,
    )

    return json(
        {
            "code": 0,
            "message": "success",
            "data": {
                "list": [
                    {
                        "id": r.id,
                        "customer_id": r.customer_id,
                        "real_amount": float(r.real_amount),
                        "bonus_amount": float(r.bonus_amount),
                        "total_amount": float(r.real_amount + r.bonus_amount),
                        "operator_id": r.operator_id,
                        "payment_proof": r.payment_proof,
                        "remark": r.remark,
                        "created_at": r.created_at.isoformat()
                        if r.created_at
                        else None,
                    }
                    for r in records
                ],
                "total": total,
                "page": page,
                "page_size": page_size,
            },
        }
    )


@billing_bp.get("/consumption-records")
@auth_required
@require_permission("billing:view")
async def get_consumption_records(request: Request):
    """获取消费记录列表"""
    db: AsyncSession = request.ctx.db_session
    from ..models.billing import ConsumptionRecord
    from sqlalchemy import select

    customer_id = (
        int(request.args.get("customer_id"))
        if request.args.get("customer_id")
        else None
    )
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 20))

    stmt = select(ConsumptionRecord).where(ConsumptionRecord.deleted_at.is_(None))

    if customer_id:
        stmt = stmt.where(ConsumptionRecord.customer_id == customer_id)

    # 总数
    from sqlalchemy import func

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await db.execute(count_stmt)).scalar()

    # 分页
    stmt = stmt.order_by(ConsumptionRecord.created_at.desc())
    stmt = stmt.offset((page - 1) * page_size).limit(page_size)

    result = await db.execute(stmt)
    records = result.scalars().all()

    return json(
        {
            "code": 0,
            "message": "success",
            "data": {
                "list": [
                    {
                        "id": r.id,
                        "customer_id": r.customer_id,
                        "invoice_id": r.invoice_id,
                        "amount": float(r.amount),
                        "bonus_used": float(r.bonus_used) if r.bonus_used else 0,
                        "real_used": float(r.real_used) if r.real_used else 0,
                        "balance_after": float(r.balance_after)
                        if r.balance_after
                        else 0,
                        "consumed_at": r.created_at.isoformat()
                        if r.created_at
                        else None,
                    }
                    for r in records
                ],
                "total": total,
                "page": page,
                "page_size": page_size,
            },
        }
    )


# ==================== 定价规则 ====================


@billing_bp.get("/pricing-rules")
@auth_required
@require_permission("billing:view")
async def get_pricing_rules(request: Request):
    """获取定价规则列表"""
    db: AsyncSession = request.ctx.db_session
    pricing_service = PricingService(db)

    customer_id = (
        int(request.args.get("customer_id"))
        if request.args.get("customer_id")
        else None
    )
    device_type = request.args.get("device_type")
    pricing_type = request.args.get("pricing_type")

    rules = await pricing_service.get_pricing_rules(
        customer_id=customer_id,
        device_type=device_type,
        pricing_type=pricing_type,
    )

    return json(
        {
            "code": 0,
            "message": "success",
            "data": [
                {
                    "id": r.id,
                    "customer_id": r.customer_id,
                    "device_type": r.device_type,
                    "pricing_type": r.pricing_type,
                    "unit_price": float(r.unit_price) if r.unit_price else None,
                    "tiers": r.tiers,
                    "package_type": r.package_type,
                    "package_limits": r.package_limits,
                    "effective_date": r.effective_date.isoformat()
                    if r.effective_date
                    else None,
                    "expiry_date": r.expiry_date.isoformat() if r.expiry_date else None,
                }
                for r in rules
            ],
        }
    )


@billing_bp.post("/pricing-rules")
@auth_required
@require_permission("billing:edit")
async def create_pricing_rule(request: Request):
    """
    创建定价规则

    Body:
    {
        "customer_id": 1,
        "device_type": "X",
        "pricing_type": "fixed",
        "unit_price": 10.00,
        "effective_date": "2026-04-01",
        "expiry_date": "2026-12-31"
    }
    """
    db: AsyncSession = request.ctx.db_session
    data = request.json
    user = get_current_user(request)

    pricing_service = PricingService(db)
    data["created_by"] = user["user_id"] if user else 1

    # 日期转换
    if "effective_date" in data and isinstance(data["effective_date"], str):
        data["effective_date"] = date.fromisoformat(data["effective_date"])
    if "expiry_date" in data and isinstance(data["expiry_date"], str):
        data["expiry_date"] = date.fromisoformat(data["expiry_date"])

    rule = await pricing_service.create_pricing_rule(data)

    # 定价规则变更后清除相关缓存
    await cache_service.invalidate_billing_cache()

    return json(
        {
            "code": 0,
            "message": "创建成功",
            "data": {
                "id": rule.id,
                "device_type": rule.device_type,
                "pricing_type": rule.pricing_type,
                "unit_price": float(rule.unit_price) if rule.unit_price else None,
            },
        },
        status=201,
    )


@billing_bp.put("/pricing-rules/<rule_id:int>")
@auth_required
@require_permission("billing:edit")
async def update_pricing_rule(request: Request, rule_id: int):
    """更新定价规则"""
    db: AsyncSession = request.ctx.db_session
    data = request.json

    pricing_service = PricingService(db)

    # 日期转换
    if "effective_date" in data and isinstance(data["effective_date"], str):
        data["effective_date"] = date.fromisoformat(data["effective_date"])
    if "expiry_date" in data and isinstance(data["expiry_date"], str):
        data["expiry_date"] = date.fromisoformat(data["expiry_date"])

    rule = await pricing_service.update_pricing_rule(rule_id, data)

    if not rule:
        return json({"code": 40401, "message": "规则不存在"}, status=404)

    # 定价规则变更后清除相关缓存
    await cache_service.invalidate_billing_cache()

    return json(
        {
            "code": 0,
            "message": "更新成功",
            "data": {
                "id": rule.id,
                "device_type": rule.device_type,
                "pricing_type": rule.pricing_type,
            },
        }
    )


@billing_bp.delete("/pricing-rules/<rule_id:int>")
@auth_required
@require_permission("billing:delete")
async def delete_pricing_rule(request: Request, rule_id: int):
    """删除定价规则"""
    db: AsyncSession = request.ctx.db_session
    pricing_service = PricingService(db)

    success = await pricing_service.delete_pricing_rule(rule_id)

    if not success:
        return json({"code": 40401, "message": "规则不存在"}, status=404)

    # 定价规则变更后清除相关缓存
    await cache_service.invalidate_billing_cache()

    return json({"code": 0, "message": "删除成功"})


# ==================== 结算单管理 ====================


@billing_bp.get("/invoices")
@auth_required
@require_permission("billing:view")
async def get_invoices(request: Request):
    """获取结算单列表"""
    db: AsyncSession = request.ctx.db_session
    invoice_service = InvoiceService(db)

    customer_id = (
        int(request.args.get("customer_id"))
        if request.args.get("customer_id")
        else None
    )
    status = request.args.get("status")
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 20))

    invoices, total = await invoice_service.get_invoices(
        customer_id=customer_id,
        status=status,
        page=page,
        page_size=page_size,
    )

    return json(
        {
            "code": 0,
            "message": "success",
            "data": {
                "list": [
                    {
                        "id": i.id,
                        "invoice_no": i.invoice_no,
                        "customer_id": i.customer_id,
                        "period_start": i.period_start.isoformat()
                        if i.period_start
                        else None,
                        "period_end": i.period_end.isoformat()
                        if i.period_end
                        else None,
                        "total_amount": float(i.total_amount),
                        "discount_amount": float(i.discount_amount)
                        if i.discount_amount
                        else 0,
                        "final_amount": float(
                            i.total_amount - (i.discount_amount or 0)
                        ),
                        "status": i.status,
                        "is_auto_generated": i.is_auto_generated,
                        "created_at": i.created_at.isoformat()
                        if i.created_at
                        else None,
                    }
                    for i in invoices
                ],
                "total": total,
                "page": page,
                "page_size": page_size,
            },
        }
    )


@billing_bp.get("/invoices/<invoice_id:int>")
@auth_required
@require_permission("billing:view")
async def get_invoice(request: Request, invoice_id: int):
    """获取结算单详情"""
    db: AsyncSession = request.ctx.db_session
    invoice_service = InvoiceService(db)

    invoice = await invoice_service.get_invoice_by_id(invoice_id)

    if not invoice:
        return json({"code": 40401, "message": "结算单不存在"}, status=404)

    return json(
        {
            "code": 0,
            "message": "success",
            "data": {
                "id": invoice.id,
                "invoice_no": invoice.invoice_no,
                "customer_id": invoice.customer_id,
                "period_start": invoice.period_start.isoformat()
                if invoice.period_start
                else None,
                "period_end": invoice.period_end.isoformat()
                if invoice.period_end
                else None,
                "total_amount": float(invoice.total_amount),
                "discount_amount": float(invoice.discount_amount)
                if invoice.discount_amount
                else 0,
                "discount_reason": invoice.discount_reason,
                "final_amount": float(
                    invoice.total_amount - (invoice.discount_amount or 0)
                ),
                "status": invoice.status,
                "items": [
                    {
                        "id": item.id,
                        "device_type": item.device_type,
                        "layer_type": item.layer_type,
                        "quantity": float(item.quantity),
                        "unit_price": float(item.unit_price),
                        "subtotal": float(item.quantity * item.unit_price),
                    }
                    for item in invoice.items
                ],
                "approver_id": invoice.approver_id,
                "approved_at": invoice.approved_at,
                "customer_confirmed_at": invoice.customer_confirmed_at,
                "paid_at": invoice.paid_at,
                "completed_at": invoice.completed_at,
            },
        }
    )


@billing_bp.post("/invoices/generate")
@auth_required
@require_permission("billing:edit")
async def generate_invoice(request: Request):
    """
    生成结算单

    Body:
    {
        "customer_id": 1,
        "period_start": "2026-03-01",
        "period_end": "2026-03-31",
        "items": [
            {"device_type": "X", "layer_type": "single", "quantity": 100, "unit_price": 10}
        ]
    }
    """
    db: AsyncSession = request.ctx.db_session
    data = request.json
    user = get_current_user(request)

    invoice_service = InvoiceService(db)

    # 日期转换
    period_start = date.fromisoformat(data["period_start"])
    period_end = date.fromisoformat(data["period_end"])
    items = data.get("items", [])

    if not items:
        return json({"code": 40001, "message": "结算项不能为空"}, status=400)

    invoice = await invoice_service.generate_invoice(
        customer_id=data["customer_id"],
        period_start=period_start,
        period_end=period_end,
        items=items,
        created_by=user["user_id"] if user else 1,
    )

    # 结算单生成后清除相关缓存
    await cache_service.invalidate_billing_cache()

    return json(
        {
            "code": 0,
            "message": "生成成功",
            "data": {
                "id": invoice.id,
                "invoice_no": invoice.invoice_no,
                "total_amount": float(invoice.total_amount),
            },
        },
        status=201,
    )


@billing_bp.put("/invoices/<invoice_id:int>/discount")
@auth_required
@require_permission("billing:edit")
async def apply_discount(request: Request, invoice_id: int):
    """
    应用减免

    Body:
    {
        "discount_amount": 500.00,
        "discount_reason": "大客户优惠",
        "discount_attachment": "/uploads/proof.xlsx"
    }
    """
    db: AsyncSession = request.ctx.db_session
    data = request.json

    invoice_service = InvoiceService(db)

    success, message = await invoice_service.apply_discount(
        invoice_id=invoice_id,
        discount_amount=Decimal(str(data.get("discount_amount", 0))),
        discount_reason=data.get("discount_reason", ""),
        discount_attachment=data.get("discount_attachment"),
    )

    if not success:
        return json({"code": 40001, "message": message}, status=400)

    # 结算单减免后清除相关缓存
    await cache_service.invalidate_billing_cache()

    return json({"code": 0, "message": message})


@billing_bp.post("/invoices/<invoice_id:int>/submit")
@auth_required
@require_permission("billing:edit")
async def submit_invoice(request: Request, invoice_id: int):
    """提交结算单（商务确认）"""
    db: AsyncSession = request.ctx.db_session
    user = get_current_user(request)
    invoice_service = InvoiceService(db)

    success, message = await invoice_service.submit_invoice(
        invoice_id=invoice_id,
        approver_id=user["user_id"] if user else 1,
    )

    if not success:
        return json({"code": 40001, "message": message}, status=400)

    # 结算单提交后清除相关缓存
    await cache_service.invalidate_billing_cache()

    return json({"code": 0, "message": message})


@billing_bp.post("/invoices/<invoice_id:int>/confirm")
@auth_required
@require_permission("billing:edit")
async def confirm_invoice(request: Request, invoice_id: int):
    """客户确认结算单"""
    db: AsyncSession = request.ctx.db_session
    invoice_service = InvoiceService(db)

    success, message = await invoice_service.confirm_invoice(invoice_id=invoice_id)

    if not success:
        return json({"code": 40001, "message": message}, status=400)

    # 结算单确认后清除相关缓存
    await cache_service.invalidate_billing_cache()

    return json({"code": 0, "message": message})


@billing_bp.post("/invoices/<invoice_id:int>/pay")
@auth_required
@require_permission("billing:edit")
async def pay_invoice(request: Request, invoice_id: int):
    """确认付款"""
    db: AsyncSession = request.ctx.db_session
    data = request.json
    invoice_service = InvoiceService(db)

    success, message = await invoice_service.pay_invoice(
        invoice_id=invoice_id,
        payment_proof=data.get("payment_proof"),
    )

    if not success:
        return json({"code": 40001, "message": message}, status=400)

    # 结算单付款后清除相关缓存
    await cache_service.invalidate_billing_cache()

    return json({"code": 0, "message": message})


@billing_bp.post("/invoices/<invoice_id:int>/complete")
@auth_required
@require_permission("billing:edit")
async def complete_invoice(request: Request, invoice_id: int):
    """完成结算（扣款）"""
    db: AsyncSession = request.ctx.db_session
    invoice_service = InvoiceService(db)

    success, message = await invoice_service.complete_invoice(invoice_id=invoice_id)

    if not success:
        return json({"code": 40001, "message": message}, status=400)

    # 结算单完成后清除相关缓存
    await cache_service.invalidate_billing_cache()

    return json({"code": 0, "message": message})


@billing_bp.delete("/invoices/<invoice_id:int>")
@auth_required
@require_permission("billing:delete")
async def delete_invoice(request: Request, invoice_id: int):
    """删除结算单"""
    db: AsyncSession = request.ctx.db_session
    invoice_service = InvoiceService(db)

    success = await invoice_service.delete_invoice(invoice_id)

    if not success:
        return json({"code": 40401, "message": "结算单不存在"}, status=404)

    # 结算单删除后清除相关缓存
    await cache_service.invalidate_billing_cache()

    return json({"code": 0, "message": "删除成功"})


@billing_bp.get("/invoices/export")
@auth_required
@require_permission("billing:view")
async def export_invoices(request: Request):
    """
    导出结算单为 Excel 文件

    查询参数:
    - customer_id: 客户 ID（可选）
    - status: 结算单状态（可选）
    - start_date: 开始日期（可选，格式：YYYY-MM-DD）
    - end_date: 结束日期（可选，格式：YYYY-MM-DD）

    响应:
    - Excel 文件下载
    """
    from ..models.billing import Invoice, InvoiceStatus
    from ..models.customers import Customer
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload

    # 获取数据库会话
    db: AsyncSession = request.ctx.db_session

    # 获取查询参数
    customer_id = (
        int(request.args.get("customer_id"))
        if request.args.get("customer_id")
        else None
    )
    status = request.args.get("status")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    # 构建基础查询
    base_stmt = (
        select(Invoice)
        .join(Customer, Invoice.customer_id == Customer.id)
        .options(selectinload(Invoice.customer))
        .where(
            Invoice.deleted_at.is_(None),
            Customer.deleted_at.is_(None),
        )
    )

    # 应用筛选条件
    if customer_id:
        base_stmt = base_stmt.where(Invoice.customer_id == customer_id)

    if status:
        # 验证状态值
        valid_statuses = [s.value for s in InvoiceStatus]
        if status not in valid_statuses:
            return json(
                {
                    "code": 40001,
                    "message": f"无效的状态值，有效值：{', '.join(valid_statuses)}",
                },
                status=400,
            )
        base_stmt = base_stmt.where(Invoice.status == status)

    if start_date:
        try:
            start = date.fromisoformat(start_date)
            base_stmt = base_stmt.where(Invoice.period_start >= start)
        except ValueError:
            return json(
                {"code": 40001, "message": "开始日期格式错误，应为 YYYY-MM-DD"},
                status=400,
            )

    if end_date:
        try:
            end = date.fromisoformat(end_date)
            base_stmt = base_stmt.where(Invoice.period_end <= end)
        except ValueError:
            return json(
                {"code": 40001, "message": "结束日期格式错误，应为 YYYY-MM-DD"},
                status=400,
            )

    # 执行查询
    result = await db.execute(base_stmt.order_by(Invoice.created_at.desc()))
    invoices = result.scalars().all()

    if not invoices:
        return json(
            {"code": 40002, "message": "没有找到符合条件的结算单"},
            status=400,
        )

    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "结算单导出"

    # 定义样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="1F2937", end_color="1F2937", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 表头
    headers = [
        "结算单号",
        "客户名称",
        "周期开始",
        "周期结束",
        "总金额",
        "折扣金额",
        "最终金额",
        "状态",
        "创建时间",
    ]

    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 设置列宽
    column_widths = [20, 30, 12, 12, 12, 12, 12, 18, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # 写入数据
    status_map = {
        "draft": "草稿",
        "pending_customer": "待客户确认",
        "customer_confirmed": "客户已确认",
        "paid": "已付款",
        "completed": "已完成",
        "cancelled": "已取消",
    }

    for row_num, invoice in enumerate(invoices, 2):
        # 获取客户名称（通过关联或查询）
        customer_name = (
            invoice.customer.name
            if hasattr(invoice, "customer") and invoice.customer
            else "未知"
        )

        row_data = [
            invoice.invoice_no,
            customer_name,
            invoice.period_start.isoformat() if invoice.period_start else "",
            invoice.period_end.isoformat() if invoice.period_end else "",
            float(invoice.total_amount),
            float(invoice.discount_amount) if invoice.discount_amount else 0,
            float(invoice.total_amount - (invoice.discount_amount or 0)),
            status_map.get(invoice.status, invoice.status),
            invoice.created_at.isoformat() if invoice.created_at else "",
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # 添加统计信息行
    total_row = len(invoices) + 2
    ws.cell(row=total_row, column=1, value=f"共 {len(invoices)} 条记录")
    ws.cell(
        row=total_row,
        column=5,
        value=f"总金额：{sum(float(inv.total_amount) for inv in invoices):.2f}",
    )
    ws.cell(
        row=total_row,
        column=7,
        value=f"最终总额：{sum(float(inv.total_amount - (inv.discount_amount or 0)) for inv in invoices):.2f}",
    )

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"结算单导出_{timestamp}.xlsx"

    # 返回文件
    return await response_file(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )
