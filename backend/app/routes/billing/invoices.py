"""发票管理路由 — 生成、审批、支付、导出"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sanic.request import Request
from sanic.response import file as response_file
from sanic.response import json
from sqlalchemy.ext.asyncio import AsyncSession

from ...cache.base import cache_service
from ...middleware.auth import auth_required, get_current_user, require_permission
from ...repository import InvoiceRepository, PricingRepository
from ...services.billing import InvoiceService
from ...utils.audit_helpers import create_audit_entry
from . import billing_bp


@billing_bp.get("/invoices")
@auth_required
@require_permission("billing:view")
async def get_invoices(request: Request):
    """获取结算单列表"""
    db: AsyncSession = request.ctx.db_session
    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    customer_id = int(request.args.get("customer_id")) if request.args.get("customer_id") else None
    keyword = request.args.get("keyword")  # 客户名称模糊搜索
    status = request.args.get("status")
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 20))

    # 排序参数
    sort_by = request.args.get("sort_by", "")
    sort_order = request.args.get("sort_order", "desc")
    if sort_order not in ("asc", "desc"):
        sort_order = "desc"

    invoices, total = await invoice_service.get_invoices(
        customer_id=customer_id,
        keyword=keyword,
        status=status,
        page=page,
        page_size=page_size,
        sort_by=sort_by,
        sort_order=sort_order,
    )

    return json(
        {
            "code": 0,
            "message": "success",
            "data": {
                "list": [
                    {
                        "id": i.id,
                        "invoice_no": i.invoice_no,
                        "customer_id": i.customer_id,
                        "customer_name": i.customer.name if i.customer else None,
                        # 客户指定的运营/销售经理（前端据此判断按钮是否可点击）
                        "customer_manager_id": i.customer.manager_id if i.customer else None,
                        "customer_sales_manager_id": i.customer.sales_manager_id
                        if i.customer
                        else None,
                        "period_start": i.period_start.isoformat() if i.period_start else None,  # pyright: ignore[reportGeneralTypeIssues]
                        "period_end": i.period_end.isoformat() if i.period_end else None,  # pyright: ignore[reportGeneralTypeIssues]
                        "total_amount": float(i.total_amount),  # pyright: ignore[reportArgumentType]
                        "discount_amount": float(i.discount_amount) if i.discount_amount else 0,  # pyright: ignore[reportArgumentType, reportGeneralTypeIssues]
                        "final_amount": float(i.total_amount - (i.discount_amount or 0)),  # pyright: ignore[reportArgumentType]
                        "status": i.status,
                        "is_auto_generated": i.is_auto_generated,
                        "created_at": i.created_at.isoformat() if i.created_at else None,  # pyright: ignore[reportGeneralTypeIssues]
                    }
                    for i in invoices
                ],
                "total": total,
                "page": page,
                "page_size": page_size,
            },
        }
    )


@billing_bp.get("/invoices/<invoice_id:int>")
@auth_required
@require_permission("billing:view")
async def get_invoice(request: Request, invoice_id: int):
    """获取结算单详情"""
    db: AsyncSession = request.ctx.db_session
    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    invoice = await invoice_service.get_invoice_by_id(invoice_id)

    if not invoice:
        return json({"code": 40401, "message": "结算单不存在"}, status=404)

    # 批量查询操作人姓名
    from ...models.users import User

    operator_ids = {
        invoice.created_by,
        invoice.approver_id,
        invoice.ops_confirmed_by,
        invoice.sales_confirmed_by,
        invoice.customer_confirmed_by,
        invoice.completed_by,
        invoice.cancelled_by,
    }
    operator_ids.discard(None)

    operator_names: dict[int, str] = {}
    if operator_ids:
        from sqlalchemy import select as sa_select

        user_result = await db.execute(
            sa_select(User.id, User.real_name, User.username).where(User.id.in_(operator_ids))
        )
        for row in user_result:
            operator_names[row.id] = row.real_name or row.username

    def resolve_name(uid):
        return operator_names.get(uid) if uid else None

    return json(
        {
            "code": 0,
            "message": "success",
            "data": {
                "id": invoice.id,
                "invoice_no": invoice.invoice_no,
                "customer_id": invoice.customer_id,
                "customer_name": invoice.customer.name if invoice.customer else None,
                # 客户指定的运营/销售经理（前端据此判断按钮是否可点击）
                "customer_manager_id": invoice.customer.manager_id if invoice.customer else None,
                "customer_sales_manager_id": invoice.customer.sales_manager_id
                if invoice.customer
                else None,
                "period_start": invoice.period_start.isoformat() if invoice.period_start else None,  # pyright: ignore[reportGeneralTypeIssues]
                "period_end": invoice.period_end.isoformat() if invoice.period_end else None,  # pyright: ignore[reportGeneralTypeIssues]
                "total_amount": float(invoice.total_amount),  # pyright: ignore[reportArgumentType]
                "discount_amount": float(invoice.discount_amount) if invoice.discount_amount else 0,  # pyright: ignore[reportArgumentType, reportGeneralTypeIssues]
                "discount_reason": invoice.discount_reason,
                "final_amount": float(invoice.total_amount - (invoice.discount_amount or 0)),  # pyright: ignore[reportArgumentType]
                "status": invoice.status,
                "items": [
                    {
                        "id": item.id,
                        "device_type": item.device_type,
                        "layer_type": item.layer_type,
                        "quantity": float(item.quantity),
                        "unit_price": float(item.unit_price),
                        "subtotal": float(item.quantity * item.unit_price),
                    }
                    for item in invoice.items
                ],
                "approver_id": invoice.approver_id,
                "approver_name": resolve_name(invoice.approver_id),
                "approved_at": invoice.approved_at,
                "discount_applied_at": invoice.discount_applied_at,
                "ops_confirmed_by": invoice.ops_confirmed_by,
                "ops_confirmed_name": resolve_name(invoice.ops_confirmed_by),
                "ops_confirmed_at": invoice.ops_confirmed_at,
                "sales_confirmed_by": invoice.sales_confirmed_by,
                "sales_confirmed_name": resolve_name(invoice.sales_confirmed_by),
                "sales_confirmed_at": invoice.sales_confirmed_at,
                "customer_confirmed_at": invoice.customer_confirmed_at,
                "customer_confirmed_by": invoice.customer_confirmed_by,
                "customer_confirmed_name": resolve_name(invoice.customer_confirmed_by),
                "paid_at": invoice.paid_at,
                "completed_at": invoice.completed_at,
                "completed_by": invoice.completed_by,
                "completed_name": resolve_name(invoice.completed_by),
                "cancelled_at": invoice.cancelled_at,
                "cancelled_by": invoice.cancelled_by,
                "cancelled_name": resolve_name(invoice.cancelled_by),
                "created_by": invoice.created_by,
                "created_by_name": resolve_name(invoice.created_by),
                "created_at": invoice.created_at.isoformat() if invoice.created_at else None,  # pyright: ignore[reportGeneralTypeIssues]
            },
        }
    )


@billing_bp.post("/invoices/calculate-items")
@auth_required
@require_permission("billing:edit")
async def calculate_invoice_items(request: Request):
    """
    根据客户 + 结算周期，自动计算结算明细

    Body:
    {
        "customer_id": 1,
        "period_start": "2026-03-01",
        "period_end": "2026-03-31"
    }

    Returns:
    {
        "code": 0,
        "message": "计算成功",
        "data": {
            "items": [
                {"device_type": "N", "layer_type": "single", "quantity": 1234, "unit_price": 10.0, "subtotal": 12340.0}
            ],
            "total_amount": 12340.0
        }
    }
    """
    db: AsyncSession = request.ctx.db_session
    data = request.json

    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    # 日期转换
    period_start = date.fromisoformat(data["period_start"])
    period_end = date.fromisoformat(data["period_end"])

    # 调用服务层计算
    items, total_amount = await invoice_service.calculate_items_from_rules(
        customer_id=data["customer_id"],
        period_start=period_start,
        period_end=period_end,
    )

    if not items:
        return json(
            {
                "code": 40002,
                "message": "该客户在结算周期内无用量数据或无匹配的计费规则",
            },
            status=400,
        )

    # 格式化返回数据
    formatted_items = [
        {
            "device_type": item["device_type"],
            "layer_type": item["layer_type"],
            "quantity": float(item["quantity"]),
            "unit_price": float(item["unit_price"]),
            "subtotal": float(item["subtotal"]),
        }
        for item in items
    ]

    return json(
        {
            "code": 0,
            "message": "计算成功",
            "data": {
                "items": formatted_items,
                "total_amount": float(total_amount),
            },
        }
    )


@billing_bp.post("/invoices/preview-batch")
@auth_required
@require_permission("billing:view")
async def preview_batch_invoices(request: Request):
    """预览批量生成结算单的匹配客户列表

    Body:
    {
        "pricing_type": "fixed",              # 计费类型（可选）
        "industry_type_ids": [1, 2, 3],       # 行业类型 ID 列表（可选）
        "scale_levels": ["S", "A"],           # 规模等级列表（可选）
        "consume_levels": ["C1", "C2"],       # 消费等级列表（可选）
        "is_real_estate": true                # 是否房产客户（可选）
    }

    Returns:
    {
        "code": 0,
        "data": {
            "total": 5,
            "customers": [
                {"id": 1, "name": "客户A", "manager_id": 3, "sales_manager_id": 5, "has_manager": true, "has_sales_manager": true}
            ]
        }
    }
    """
    db: AsyncSession = request.ctx.db_session
    data = request.json or {}

    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    customers = await invoice_service.get_customers_for_batch(
        pricing_type=data.get("pricing_type"),
        industry_type_ids=data.get("industry_type_ids"),
        scale_levels=data.get("scale_levels"),
        consume_levels=data.get("consume_levels"),
        is_real_estate=data.get("is_real_estate"),
    )

    # 标记是否已指定经理
    result_list = [
        {
            "id": c["id"],
            "name": c["name"],
            "manager_id": c["manager_id"],
            "sales_manager_id": c["sales_manager_id"],
            "has_manager": bool(c["manager_id"]),
            "has_sales_manager": bool(c["sales_manager_id"]),
        }
        for c in customers
    ]

    return json(
        {
            "code": 0,
            "message": "success",
            "data": {"total": len(result_list), "customers": result_list},
        }
    )


@billing_bp.post("/invoices/generate-batch")
@auth_required
@require_permission("billing:edit")
async def generate_invoices_batch(request: Request):
    """按计费类型批量生成结算单

    为每个匹配的客户独立生成一张结算单（状态 draft）。

    Body:
    {
        "pricing_type": "fixed",
        "industry_type_ids": [1, 2, 3],
        "scale_levels": ["S", "A"],
        "consume_levels": ["C1", "C2"],
        "is_real_estate": true,
        "period_start": "2026-07-01",
        "period_end": "2026-07-31"
    }

    Returns:
    {
        "code": 0,
        "message": "批量生成完成",
        "data": {
            "success_count": 3,
            "generated": [...],
            "skipped": [...]
        }
    }
    """
    db: AsyncSession = request.ctx.db_session
    data = request.json
    user = get_current_user(request)

    period_start = date.fromisoformat(data["period_start"])
    period_end = date.fromisoformat(data["period_end"])

    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    result = await invoice_service.generate_invoices_batch(
        pricing_type=data.get("pricing_type"),
        industry_type_ids=data.get("industry_type_ids"),
        scale_levels=data.get("scale_levels"),
        consume_levels=data.get("consume_levels"),
        is_real_estate=data.get("is_real_estate"),
        period_start=period_start,
        period_end=period_end,
        created_by=user["user_id"] if user else 1,
    )

    # 批量生成后清除缓存
    await cache_service.invalidate_billing_cache()

    # 记录审计日志
    await create_audit_entry(
        db_session=db,
        user_id=user.get("user_id") if user else None,
        action="generate_batch",
        module="billing",
        record_id=0,
        record_type="invoice",
        changes={
            "after": {
                "success_count": result["success_count"],
                "skipped_count": len(result["skipped"]),
            }
        },
        operation_type="batch",
        ip_address=request.headers.get(
            "x-real-ip", request.headers.get("x-forwarded-for", request.ip)
        ),
        auto_commit=True,
    )

    return json(
        {
            "code": 0,
            "message": f"批量生成完成：成功 {result['success_count']} 个，跳过 {len(result['skipped'])} 个",
            "data": result,
        }
    )


@billing_bp.post("/invoices/generate")
@auth_required
@require_permission("billing:edit")
async def generate_invoice(request: Request):
    """
    生成结算单

    Body:
    {
        "customer_id": 1,
        "period_start": "2026-03-01",
        "period_end": "2026-03-31",
        "items": [
            {"device_type": "X", "layer_type": "single", "quantity": 100, "unit_price": 10}
        ]
    }

    注意：如果未提供 items，系统会自动根据客户的计费规则和用量数据生成。
    """
    db: AsyncSession = request.ctx.db_session
    data = request.json
    user = get_current_user(request)

    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    # 日期转换
    period_start = date.fromisoformat(data["period_start"])
    period_end = date.fromisoformat(data["period_end"])

    # 区分"未提供 items"和"items 为空列表"
    if "items" not in data:
        # 未提供 items，自动根据计费规则 + 用量计算
        items, _ = await invoice_service.calculate_items_from_rules(
            customer_id=data["customer_id"],
            period_start=period_start,
            period_end=period_end,
        )
        if not items:
            return json(
                {
                    "code": 40002,
                    "message": "该客户在结算周期内无用量数据或无匹配的计费规则",
                },
                status=400,
            )
    elif not data["items"]:
        # 显式传入空列表，拒绝
        return json(
            {
                "code": 40001,
                "message": "结算项不能为空",
            },
            status=400,
        )
    else:
        items = data["items"]
    invoice = await invoice_service.generate_invoice(
        customer_id=data["customer_id"],
        period_start=period_start,
        period_end=period_end,
        items=items,
        created_by=user["user_id"] if user else 1,
    )

    # 结算单生成后清除相关缓存
    await cache_service.invalidate_billing_cache()

    return json(
        {
            "code": 0,
            "message": "生成成功",
            "data": {
                "id": invoice.id,
                "invoice_no": invoice.invoice_no,
                "total_amount": float(invoice.total_amount),  # pyright: ignore[reportArgumentType]
            },
        },
        status=201,
    )


@billing_bp.put("/invoices/<invoice_id:int>/discount")
@auth_required
@require_permission("billing:edit")
async def apply_discount(request: Request, invoice_id: int):
    """
    应用减免

    Body:
    {
        "discount_amount": 500.00,
        "discount_reason": "大客户优惠",
        "discount_attachment": "/uploads/proof.xlsx"
    }
    """
    db: AsyncSession = request.ctx.db_session
    data = request.json

    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    success, message = await invoice_service.apply_discount(
        invoice_id=invoice_id,
        discount_amount=Decimal(str(data.get("discount_amount", 0))),
        discount_reason=data.get("discount_reason", ""),
        discount_attachment=data.get("discount_attachment"),
    )

    if not success:
        return json({"code": 40001, "message": message}, status=400)

    # 结算单减免后清除相关缓存
    await cache_service.invalidate_billing_cache()

    return json({"code": 0, "message": message})


@billing_bp.post("/invoices/<invoice_id:int>/submit")
@auth_required
@require_permission("billing:edit")
async def submit_invoice(request: Request, invoice_id: int):
    """提交结算单（商务确认）"""
    db: AsyncSession = request.ctx.db_session
    user = get_current_user(request)
    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    # 获取提交前状态
    invoice_before = await invoice_service.get_invoice_by_id(invoice_id)
    status_before = invoice_before.status if invoice_before else None

    success, message = await invoice_service.submit_invoice(
        invoice_id=invoice_id,
        approver_id=user["user_id"] if user else 1,
    )

    if not success:
        return json({"code": 40001, "message": message}, status=400)

    # 获取提交后状态
    invoice_after = await invoice_service.get_invoice_by_id(invoice_id)

    # 结算单提交后清除相关缓存
    await cache_service.invalidate_billing_cache()

    # 记录审计日志
    await create_audit_entry(
        db_session=db,
        user_id=user.get("user_id") if user else None,
        action="submit",
        module="billing",
        record_id=invoice_id,
        record_type="invoice",
        changes={
            "before": {"status": status_before},
            "after": {"status": invoice_after.status if invoice_after else None},
        },
        operation_type="standard",
        ip_address=request.headers.get(
            "x-real-ip", request.headers.get("x-forwarded-for", request.ip)
        ),
        auto_commit=True,
    )

    return json({"code": 0, "message": message})


@billing_bp.post("/invoices/<invoice_id:int>/confirm")
@auth_required
@require_permission("billing:confirm")
async def confirm_invoice(request: Request, invoice_id: int):
    """客户确认结算单（第三步，线下确认后线上录入）

    确认后系统自动执行余额扣款，成功则进入 completed。
    """
    db: AsyncSession = request.ctx.db_session
    user = get_current_user(request)
    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    # 获取确认前状态
    invoice_before = await invoice_service.get_invoice_by_id(invoice_id)
    status_before = invoice_before.status if invoice_before else None

    success, message = await invoice_service.confirm_invoice(
        invoice_id=invoice_id,
        user_id=user["user_id"] if user else 1,
    )

    if not success:
        return json({"code": 40001, "message": message}, status=400)

    # 获取确认后状态
    invoice_after = await invoice_service.get_invoice_by_id(invoice_id)

    # 结算单确认后清除相关缓存
    await cache_service.invalidate_billing_cache()

    # 记录审计日志
    await create_audit_entry(
        db_session=db,
        user_id=user.get("user_id") if user else None,
        action="confirm",
        module="billing",
        record_id=invoice_id,
        record_type="invoice",
        changes={
            "before": {"status": status_before},
            "after": {"status": invoice_after.status if invoice_after else None},
        },
        operation_type="standard",
        ip_address=request.headers.get(
            "x-real-ip", request.headers.get("x-forwarded-for", request.ip)
        ),
        auto_commit=True,
    )

    return json({"code": 0, "message": message})


@billing_bp.post("/invoices/<invoice_id:int>/confirm-ops")
@auth_required
@require_permission("billing:ops_approve")
async def confirm_ops(request: Request, invoice_id: int):
    """运营经理确认结算单（第一步）"""
    db: AsyncSession = request.ctx.db_session
    user = get_current_user(request)
    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    invoice_before = await invoice_service.get_invoice_by_id(invoice_id)
    status_before = invoice_before.status if invoice_before else None

    success, message = await invoice_service.confirm_ops(
        invoice_id=invoice_id,
        user_id=user["user_id"] if user else 1,
    )

    if not success:
        return json({"code": 40001, "message": message}, status=400)

    invoice_after = await invoice_service.get_invoice_by_id(invoice_id)
    await cache_service.invalidate_billing_cache()

    await create_audit_entry(
        db_session=db,
        user_id=user.get("user_id") if user else None,
        action="confirm_ops",
        module="billing",
        record_id=invoice_id,
        record_type="invoice",
        changes={
            "before": {"status": status_before},
            "after": {"status": invoice_after.status if invoice_after else None},
        },
        operation_type="standard",
        ip_address=request.headers.get(
            "x-real-ip", request.headers.get("x-forwarded-for", request.ip)
        ),
        auto_commit=True,
    )

    return json({"code": 0, "message": message})


@billing_bp.post("/invoices/<invoice_id:int>/confirm-sales")
@auth_required
@require_permission("billing:sales_approve")
async def confirm_sales(request: Request, invoice_id: int):
    """销售经理确认结算单（第二步）"""
    db: AsyncSession = request.ctx.db_session
    user = get_current_user(request)
    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    invoice_before = await invoice_service.get_invoice_by_id(invoice_id)
    status_before = invoice_before.status if invoice_before else None

    success, message = await invoice_service.confirm_sales(
        invoice_id=invoice_id,
        user_id=user["user_id"] if user else 1,
    )

    if not success:
        return json({"code": 40001, "message": message}, status=400)

    invoice_after = await invoice_service.get_invoice_by_id(invoice_id)
    await cache_service.invalidate_billing_cache()

    await create_audit_entry(
        db_session=db,
        user_id=user.get("user_id") if user else None,
        action="confirm_sales",
        module="billing",
        record_id=invoice_id,
        record_type="invoice",
        changes={
            "before": {"status": status_before},
            "after": {"status": invoice_after.status if invoice_after else None},
        },
        operation_type="standard",
        ip_address=request.headers.get(
            "x-real-ip", request.headers.get("x-forwarded-for", request.ip)
        ),
        auto_commit=True,
    )

    return json({"code": 0, "message": message})


@billing_bp.post("/invoices/<invoice_id:int>/retry-deduction")
@auth_required
@require_permission("billing:confirm")
async def retry_deduction(request: Request, invoice_id: int):
    """重试扣款（客户确认后扣款失败时手动重试）"""
    db: AsyncSession = request.ctx.db_session
    user = get_current_user(request)
    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    invoice_before = await invoice_service.get_invoice_by_id(invoice_id)
    status_before = invoice_before.status if invoice_before else None

    success, message = await invoice_service.retry_deduction(
        invoice_id=invoice_id,
        user_id=user.get("user_id") if user else 1,  # pyright: ignore[reportArgumentType]
    )

    if not success:
        return json({"code": 40001, "message": message}, status=400)

    invoice_after = await invoice_service.get_invoice_by_id(invoice_id)
    await cache_service.invalidate_billing_cache()

    await create_audit_entry(
        db_session=db,
        user_id=user.get("user_id") if user else None,
        action="retry_deduction",
        module="billing",
        record_id=invoice_id,
        record_type="invoice",
        changes={
            "before": {"status": status_before},
            "after": {"status": invoice_after.status if invoice_after else None},
        },
        operation_type="standard",
        ip_address=request.headers.get(
            "x-real-ip", request.headers.get("x-forwarded-for", request.ip)
        ),
        auto_commit=True,
    )

    return json({"code": 0, "message": message})


@billing_bp.post("/invoices/<invoice_id:int>/pay")
@auth_required
@require_permission("billing:pay")
async def pay_invoice(request: Request, invoice_id: int):
    """确认付款"""
    db: AsyncSession = request.ctx.db_session
    user = get_current_user(request)
    data = request.json or {}
    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    # 获取付款前状态
    invoice_before = await invoice_service.get_invoice_by_id(invoice_id)
    status_before = invoice_before.status if invoice_before else None

    success, message = await invoice_service.pay_invoice(
        invoice_id=invoice_id,
        payment_proof=data.get("payment_proof"),
    )

    if not success:
        return json({"code": 40001, "message": message}, status=400)

    # 获取付款后状态
    invoice_after = await invoice_service.get_invoice_by_id(invoice_id)

    # 结算单付款后清除相关缓存
    await cache_service.invalidate_billing_cache()

    # 记录审计日志
    await create_audit_entry(
        db_session=db,
        user_id=user.get("user_id") if user else None,
        action="pay",
        module="billing",
        record_id=invoice_id,
        record_type="invoice",
        changes={
            "before": {"status": status_before, "payment_proof": None},
            "after": {
                "status": invoice_after.status if invoice_after else None,
                "payment_proof": data.get("payment_proof"),
            },
        },
        operation_type="standard",
        ip_address=request.headers.get(
            "x-real-ip", request.headers.get("x-forwarded-for", request.ip)
        ),
        auto_commit=True,
    )

    return json({"code": 0, "message": message})


@billing_bp.post("/invoices/<invoice_id:int>/complete")
@auth_required
@require_permission("billing:pay")
async def complete_invoice(request: Request, invoice_id: int):
    """完成结算（扣款）"""
    db: AsyncSession = request.ctx.db_session
    user = get_current_user(request)
    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    # 获取完成前状态
    invoice_before = await invoice_service.get_invoice_by_id(invoice_id)
    status_before = invoice_before.status if invoice_before else None

    success, message = await invoice_service.complete_invoice(
        invoice_id=invoice_id,
        user_id=user.get("user_id") if user else 1,  # pyright: ignore[reportArgumentType]
    )

    if not success:
        return json({"code": 40001, "message": message}, status=400)

    # 获取完成后状态
    invoice_after = await invoice_service.get_invoice_by_id(invoice_id)

    # 结算单完成后清除相关缓存
    await cache_service.invalidate_billing_cache()

    # 记录审计日志
    await create_audit_entry(
        db_session=db,
        user_id=user.get("user_id") if user else None,
        action="complete",
        module="billing",
        record_id=invoice_id,
        record_type="invoice",
        changes={
            "before": {"status": status_before},
            "after": {"status": invoice_after.status if invoice_after else None},
        },
        operation_type="standard",
        ip_address=request.headers.get(
            "x-real-ip", request.headers.get("x-forwarded-for", request.ip)
        ),
        auto_commit=True,
    )

    return json({"code": 0, "message": message})


@billing_bp.post("/invoices/<invoice_id:int>/cancel")
@auth_required
@require_permission("billing:edit")
async def cancel_invoice_route(request: Request, invoice_id: int):
    """取消结算单"""
    db: AsyncSession = request.ctx.db_session
    user = get_current_user(request)
    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    # 获取取消前状态
    invoice_before = await invoice_service.get_invoice_by_id(invoice_id)
    status_before = invoice_before.status if invoice_before else None

    success, message = await invoice_service.cancel_invoice(
        invoice_id=invoice_id,
        user_id=user["user_id"] if user else None,  # pyright: ignore[reportArgumentType]
    )

    if not success:
        return json({"code": 40001, "message": message}, status=400)

    # 获取取消后状态
    invoice_after = await invoice_service.get_invoice_by_id(invoice_id)

    # 结算单取消后清除相关缓存
    await cache_service.invalidate_billing_cache()

    # 记录审计日志
    await create_audit_entry(
        db_session=db,
        user_id=user.get("user_id") if user else None,
        action="cancel",
        module="billing",
        record_id=invoice_id,
        record_type="invoice",
        changes={
            "before": {"status": status_before},
            "after": {"status": invoice_after.status if invoice_after else None},
        },
        operation_type="standard",
        ip_address=request.headers.get(
            "x-real-ip", request.headers.get("x-forwarded-for", request.ip)
        ),
        auto_commit=True,
    )

    return json({"code": 0, "message": message})


@billing_bp.delete("/invoices/<invoice_id:int>")
@auth_required
@require_permission("billing:delete")
async def delete_invoice(request: Request, invoice_id: int):
    """删除结算单"""
    db: AsyncSession = request.ctx.db_session
    invoice_service = InvoiceService(InvoiceRepository(db), PricingRepository(db))

    success = await invoice_service.delete_invoice(invoice_id)

    if not success:
        return json({"code": 40401, "message": "结算单不存在"}, status=404)

    # 结算单删除后清除相关缓存
    await cache_service.invalidate_billing_cache()

    return json({"code": 0, "message": "删除成功"})


@billing_bp.get("/invoices/export")
@auth_required
@require_permission("billing:view")
async def export_invoices(request: Request):
    """
    导出结算单为 Excel 文件

    查询参数:
    - customer_id: 客户 ID（可选）
    - status: 结算单状态（可选）
    - start_date: 开始日期（可选，格式：YYYY-MM-DD）
    - end_date: 结束日期（可选，格式：YYYY-MM-DD）

    响应:
    - Excel 文件下载
    """
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload

    from ...models.billing import Invoice, InvoiceStatus
    from ...models.customers import Customer

    # 获取数据库会话
    db: AsyncSession = request.ctx.db_session

    # 获取查询参数
    customer_id = int(request.args.get("customer_id")) if request.args.get("customer_id") else None
    status = request.args.get("status")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    # 构建基础查询
    base_stmt = (
        select(Invoice)
        .join(Customer, Invoice.customer_id == Customer.id)
        .options(selectinload(Invoice.customer))
        .where(
            Invoice.deleted_at.is_(None),
            Customer.deleted_at.is_(None),
        )
    )

    # 应用筛选条件
    if customer_id:
        base_stmt = base_stmt.where(Invoice.customer_id == customer_id)

    if status:
        # 验证状态值
        valid_statuses = [s.value for s in InvoiceStatus]
        if status not in valid_statuses:
            return json(
                {
                    "code": 40001,
                    "message": f"无效的状态值，有效值：{', '.join(valid_statuses)}",
                },
                status=400,
            )
        base_stmt = base_stmt.where(Invoice.status == status)

    if start_date:
        try:
            start = date.fromisoformat(start_date)
            base_stmt = base_stmt.where(Invoice.period_start >= start)
        except ValueError:
            return json(
                {"code": 40001, "message": "开始日期格式错误，应为 YYYY-MM-DD"},
                status=400,
            )

    if end_date:
        try:
            end = date.fromisoformat(end_date)
            base_stmt = base_stmt.where(Invoice.period_end <= end)
        except ValueError:
            return json(
                {"code": 40001, "message": "结束日期格式错误，应为 YYYY-MM-DD"},
                status=400,
            )

    # 执行查询
    result = await db.execute(base_stmt.order_by(Invoice.created_at.desc()))
    invoices = result.scalars().all()

    if not invoices:
        return json(
            {"code": 40002, "message": "没有找到符合条件的结算单"},
            status=400,
        )

    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "结算单导出"  # pyright: ignore[reportOptionalMemberAccess]

    # 定义样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 表头
    headers = [
        "结算单号",
        "客户名称",
        "周期开始",
        "周期结束",
        "总金额",
        "折扣金额",
        "最终金额",
        "状态",
        "创建时间",
    ]

    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)  # pyright: ignore[reportOptionalMemberAccess]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 设置列宽
    column_widths = [20, 30, 12, 12, 12, 12, 12, 18, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width  # pyright: ignore[reportOptionalMemberAccess]

    # 写入数据
    status_map = {
        "draft": "草稿",
        "pending_customer": "待客户确认",
        "customer_confirmed": "客户已确认",
        "paid": "已付款",
        "completed": "已完成",
        "cancelled": "已取消",
    }

    for row_num, invoice in enumerate(invoices, 2):
        # 获取客户名称（通过关联或查询）
        customer_name = (
            invoice.customer.name if hasattr(invoice, "customer") and invoice.customer else "未知"
        )

        row_data = [
            invoice.invoice_no,
            customer_name,
            invoice.period_start.isoformat() if invoice.period_start else "",  # pyright: ignore[reportGeneralTypeIssues]
            invoice.period_end.isoformat() if invoice.period_end else "",  # pyright: ignore[reportGeneralTypeIssues]
            float(invoice.total_amount),  # pyright: ignore[reportArgumentType]
            float(invoice.discount_amount) if invoice.discount_amount else 0,  # pyright: ignore[reportArgumentType, reportGeneralTypeIssues]
            float(invoice.total_amount - (invoice.discount_amount or 0)),  # pyright: ignore[reportArgumentType]
            status_map.get(invoice.status, invoice.status),  # pyright: ignore[reportCallIssue, reportArgumentType]
            invoice.created_at.isoformat() if invoice.created_at else "",  # pyright: ignore[reportGeneralTypeIssues]
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)  # pyright: ignore[reportOptionalMemberAccess]
            cell.alignment = cell_alignment
            cell.border = thin_border

    # 添加统计信息行
    total_row = len(invoices) + 2
    ws.cell(row=total_row, column=1, value=f"共 {len(invoices)} 条记录")  # pyright: ignore[reportOptionalMemberAccess]
    ws.cell(  # pyright: ignore[reportOptionalMemberAccess]
        row=total_row,
        column=5,
        value=f"总金额：{sum(float(inv.total_amount) for inv in invoices):.2f}",  # pyright: ignore[reportArgumentType]
    )
    ws.cell(  # pyright: ignore[reportOptionalMemberAccess]
        row=total_row,
        column=7,
        value=f"最终总额：{sum(float(inv.total_amount - (inv.discount_amount or 0)) for inv in invoices):.2f}",  # pyright: ignore[reportArgumentType]
    )

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"结算单导出_{timestamp}.xlsx"

    # 返回文件
    return await response_file(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # pyright: ignore[reportCallIssue]
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


# ==================== 余额趋势 ====================
